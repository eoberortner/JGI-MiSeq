from openpyxl import Workbook
from openpyxl.style import Color, Fill

result_colors = {'perfect':'1F7218',
                 'almost':'68BE60',
                 'incomplete':'D1CB87',
                 'lowcov':'CCCCCC',
                 'errors':'E49999',
                 'dips':'84CCC9',
                 'nocall':'999999'
                }

text_colors =   {'perfect':'FFFFFF',
                 'almost':'FFFFFF',
                 'incomplete':'A53E3E',
                 'lowcov':'A53E3E',
                 'errors':'A53E3E',
                 'dips':'233D69',
                 'nocall':'CCCCCC'
                }

def excel_style(row, col):
    """ Convert given row and column number to an Excel-style cell name. """
    quot, rem = divmod(ord(col)-ord('A'), 26)
    return((chr(quot-1 + ord('A')) if quot else '') +
           (chr(rem + ord('A')) + str(row)))

def create_result_workbook(refs,pools,calltable,bestbets,outfile):
    wb = Workbook()
    ws = wb.active

    ws.title = "bestbets"
    ws.append(["clone","best pool","variant","univF - reverse primer","forward - univR primer"])
    row = 2
    for ref in refs:
        ws['A%d' % row] = ref
        if bestbets[ref] is not None:
            c = calltable[ref][bestbets[ref]]
            ws['B%d' % row] = bestbets[ref]
      
            ws['B%d' % row].style.fill.fill_type = Fill.FILL_SOLID
            ws['B%d' % row].style.fill.start_color.index = result_colors[c['call']]
            if c['call'] == 'almost':
                ws['D%d' % row] = c['p1']
                ws['E%d' % row] = c['p2']
                for col in ['C','D','E']:
                    ws['%s%d' % (col,row)].style.fill.fill_type = Fill.FILL_SOLID        
                    ws['%s%d' % (col,row)].style.fill.start_color.index = result_colors[c['call']]
        
        row += 1
  
    ws = wb.create_sheet()
    ws.title = 'overview'
    ws.append([""] + pools)
 
    cols = dict((p,chr(66+i)) for i,p in enumerate(pools))

    row = 2
    for ref in refs:
        ws['A%d' % row] = ref
        for p in pools:
            c = calltable[ref][p] #['call']

            ## modified 07/27/15:
            ## -- introduced the excel_style function to count like Excel counts its columns:
            ##    A, ..., Z, AA, AB, AC, ..., AZ, BA, ..., BZ, ...
            excel_cell = excel_style(row, cols[p])
            ws[excel_cell] = c['call']
            ws[excel_cell].style.fill.fill_type = Fill.FILL_SOLID         
            ws[excel_cell].style.fill.start_color.index = result_colors[c['call']]
            ws[excel_cell].style.font.color.index = text_colors[c['call']]


        row += 1
  
  
    for pool in pools:
        ws = wb.create_sheet()
        ws.title = pool
        ws.append(["clone","result","variant","univF - reverse primer","forward - univR primer"])
        row = 2
        for ref in refs:
            ws['A%d' % row] = ref
            c = calltable[ref][pool]
            ws['B%d' % row] = c['call']
            ws['B%d' % row].style.fill.fill_type = Fill.FILL_SOLID
            ws['B%d' % row].style.fill.start_color.index = result_colors[c['call']]
            ws['B%d' % row].style.font.color.index = text_colors[c['call']]
            if c['call'] == 'almost':
                ws['D%d' % row] = c['p1']
                ws['E%d' % row] = c['p2']
                for col in ['C','D','E']:
                    ws['%s%d' % (col,row)].style.fill.fill_type = Fill.FILL_SOLID        
                    ws['%s%d' % (col,row)].style.fill.start_color.index = result_colors[c['call']]
                    ws['%s%d' % (col,row)].style.font.color.index = text_colors[c['call']]
            
            row += 1
  
  
    wb.save(filename=outfile)  
